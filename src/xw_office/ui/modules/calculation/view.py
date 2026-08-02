"""Provisionen / Kalkulation module."""
from __future__ import annotations

import csv
from io import StringIO
import logging
from pathlib import Path
from typing import TYPE_CHECKING

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QDateEdit,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QSplitter,
    QStackedWidget,
    QSizePolicy,
    QTableView,
    QVBoxLayout,
    QWidget,
)
from openpyxl import Workbook

from xw_office.core.worker import BackgroundWorker
from xw_office.services.commission.service import CommissionRunResult, CommissionService
from xw_office.services.calculation.service import (
    ArticleEntry,
    CalculationService,
    calculate_royalty,
)
from xw_office.ui.widgets.data_table import DataTable

if TYPE_CHECKING:
    from xw_office.core.container import Container

logger = logging.getLogger(__name__)

_ARTICLE_HEADERS = ["Titel", "Brutto EUR", "MwSt %", "Provision %", "Netto EUR", "MwSt EUR", "Provision EUR", "Notiz"]
_PRODUCT_HEADERS = [
    "SKU",
    "Name",
    "Verkauft",
    "Storno",
    "Gutschrift",
    "Netto-Menge",
    "Netto EUR",
    "Brutto EUR",
    "Kategorien",
    "Warnung",
]
_CATEGORY_HEADERS = ["Kategorie", "Menge", "Netto EUR", "Brutto EUR", "Anteil Netto"]
_DOC_HEADERS = ["Beleg", "Datum", "Typ", "SKU", "Menge", "Netto", "Regel"]


class CalculationView(QWidget):
    """Commission workspace with grouped navigation and calculators."""

    def __init__(self, container: Container, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._container = container
        self._commission_service: CommissionService = self._container.resolve(CommissionService)
        self._articles: list[ArticleEntry] = []
        self._worker: BackgroundWorker | None = None
        self._commission_worker: BackgroundWorker | None = None
        self._export_worker: BackgroundWorker | None = None
        self._last_commission_result: CommissionRunResult | None = None
        self._active_profile_key: str = ""

        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setChildrenCollapsible(False)

        nav_wrap = QWidget()
        nav_lay = QVBoxLayout(nav_wrap)
        nav_lay.setContentsMargins(0, 0, 0, 0)
        nav_lay.setSpacing(6)
        nav_title = QLabel("Provisionen")
        nav_title.setStyleSheet("font-weight: 600;")
        nav_lay.addWidget(nav_title)

        self._nav_filter = QLineEdit()
        self._nav_filter.setPlaceholderText("Auswahl filtern...")
        self._nav_filter.textChanged.connect(self._apply_nav_filter)
        nav_lay.addWidget(self._nav_filter)

        self._nav = QListWidget()
        self._nav.setAlternatingRowColors(False)
        self._nav.currentRowChanged.connect(self._on_nav_changed)
        nav_lay.addWidget(self._nav, stretch=1)
        nav_wrap.setMinimumWidth(240)
        nav_wrap.setMaximumWidth(320)

        self._stack = QStackedWidget()
        self._commission_page_index = self._stack.addWidget(self._build_musikheroes_page())
        self._add_nav_header("Abrechnungen")
        profiles = self._commission_service.list_profiles()
        for profile in profiles:
            self._add_nav_entry(
                profile.label,
                page_index=self._commission_page_index,
                kind="profile",
                key=profile.key,
            )

        self._add_nav_header("Kalkulatoren")
        self._add_nav_entry(
            "Artikelliste",
            page_index=self._stack.addWidget(self._build_articles_tab()),
            kind="page",
            key="articles",
        )
        self._add_nav_entry(
            "Schnellrechner",
            page_index=self._stack.addWidget(self._build_calc_tab()),
            kind="page",
            key="quickcalc",
        )

        if profiles:
            self._active_profile_key = profiles[0].key
            self._commission_title.setText(f"{profiles[0].label} Abrechnung")

        splitter.addWidget(nav_wrap)
        splitter.addWidget(self._stack)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        root.addWidget(splitter)

        self._select_first_nav_page()

        self._load_articles()

    def _add_nav_header(self, text: str) -> None:
        item = QListWidgetItem(text)
        item.setFlags(Qt.ItemFlag.NoItemFlags)
        font = item.font()
        font.setBold(True)
        item.setFont(font)
        item.setData(Qt.ItemDataRole.UserRole, {"kind": "header"})
        self._nav.addItem(item)

    def _add_nav_entry(self, text: str, *, page_index: int, kind: str, key: str) -> None:
        item = QListWidgetItem(f"  {text}")
        item.setData(
            Qt.ItemDataRole.UserRole,
            {
                "kind": kind,
                "key": key,
                "page_index": page_index,
                "label": text,
            },
        )
        self._nav.addItem(item)

    def _apply_nav_filter(self, text: str) -> None:
        needle = text.strip().lower()
        for row in range(self._nav.count()):
            item = self._nav.item(row)
            payload = item.data(Qt.ItemDataRole.UserRole)
            if not isinstance(payload, dict):
                continue
            kind = str(payload.get("kind") or "")
            if kind == "header":
                item.setHidden(False)
                continue
            label = str(payload.get("label") or item.text()).strip().lower()
            item.setHidden(bool(needle) and needle not in label)

    def _select_first_nav_page(self) -> None:
        for row in range(self._nav.count()):
            item = self._nav.item(row)
            data = item.data(Qt.ItemDataRole.UserRole)
            if isinstance(data, dict) and str(data.get("kind") or "") in {"page", "profile"}:
                self._nav.setCurrentRow(row)
                return

    def _on_nav_changed(self, row: int) -> None:
        if row < 0:
            return
        item = self._nav.item(row)
        payload = item.data(Qt.ItemDataRole.UserRole)
        if not isinstance(payload, dict):
            return
        kind = str(payload.get("kind") or "")
        page_index = payload.get("page_index")
        if not isinstance(page_index, int):
            return
        self._stack.setCurrentIndex(page_index)
        if kind == "profile":
            key = str(payload.get("key") or "").strip()
            if key:
                self._active_profile_key = key
                profile = self._commission_service.get_profile(key)
                self._commission_title.setText(f"{profile.label} Abrechnung")

    # ------------------------------------------------------------------
    # Commission page: MusikHeroes
    # ------------------------------------------------------------------

    def _build_musikheroes_page(self) -> QWidget:
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(8)

        self._commission_title = QLabel("Abrechnung")
        self._commission_title.setStyleSheet("font-size: 16px; font-weight: 600;")
        lay.addWidget(self._commission_title)

        controls = QHBoxLayout()
        controls.addWidget(QLabel("Zeitraum:"))
        self._period_combo = QComboBox()
        self._period_combo.addItem("Letzter Monat", "last_month")
        self._period_combo.addItem("Letztes Quartal", "last_quarter")
        self._period_combo.addItem("Letztes Halbjahr", "last_half_year")
        self._period_combo.addItem("Letztes Jahr", "last_year")
        self._period_combo.addItem("Benutzerdefiniert", "custom")
        self._period_combo.currentIndexChanged.connect(self._on_period_changed)
        controls.addWidget(self._period_combo)

        controls.addWidget(QLabel("Stichtag:"))
        self._reference_date = QDateEdit()
        self._reference_date.setCalendarPopup(True)
        self._reference_date.setDate(self._reference_date.date().currentDate())
        controls.addWidget(self._reference_date)

        controls.addWidget(QLabel("Von:"))
        self._custom_start = QDateEdit()
        self._custom_start.setCalendarPopup(True)
        self._custom_start.setEnabled(False)
        controls.addWidget(self._custom_start)

        controls.addWidget(QLabel("Bis:"))
        self._custom_end = QDateEdit()
        self._custom_end.setCalendarPopup(True)
        self._custom_end.setEnabled(False)
        controls.addWidget(self._custom_end)
        controls.addStretch()
        lay.addLayout(controls)

        toggles = QHBoxLayout()
        self._include_cancellations = QCheckBox("Stornos einbeziehen")
        self._include_cancellations.setChecked(True)
        toggles.addWidget(self._include_cancellations)
        self._include_credit_notes = QCheckBox("Gutschriften einbeziehen")
        self._include_credit_notes.setChecked(True)
        toggles.addWidget(self._include_credit_notes)
        self._show_anomalies = QCheckBox("Problemfaelle anzeigen")
        self._show_anomalies.setChecked(True)
        toggles.addWidget(self._show_anomalies)
        toggles.addStretch()

        run_btn = QPushButton("Neu laden")
        run_btn.clicked.connect(lambda: self._run_musikheroes(use_cache=False))
        toggles.addWidget(run_btn)
        cache_btn = QPushButton("Cache verwenden")
        cache_btn.clicked.connect(lambda: self._run_musikheroes(use_cache=True))
        toggles.addWidget(cache_btn)
        export_csv_btn = QPushButton("CSV exportieren")
        export_csv_btn.clicked.connect(self._export_commission_csv)
        toggles.addWidget(export_csv_btn)
        export_xlsx_btn = QPushButton("XLSX exportieren")
        export_xlsx_btn.clicked.connect(self._export_commission_xlsx)
        toggles.addWidget(export_xlsx_btn)
        copy_btn = QPushButton("Abrechnung kopieren")
        copy_btn.clicked.connect(self._copy_commission_summary)
        toggles.addWidget(copy_btn)
        lay.addLayout(toggles)

        self._commission_status = QLabel("Noch nicht geladen")
        lay.addWidget(self._commission_status)

        kpi = QHBoxLayout()
        self._kpi_qty = QLabel("Menge: -")
        self._kpi_net = QLabel("Netto: -")
        self._kpi_gross = QLabel("Brutto: -")
        self._kpi_corr = QLabel("Korrekturen: -")
        self._kpi_docs = QLabel("Belege: -")
        self._kpi_anomalies = QLabel("Problemfaelle: -")
        for label in (
            self._kpi_qty,
            self._kpi_net,
            self._kpi_gross,
            self._kpi_corr,
            self._kpi_docs,
            self._kpi_anomalies,
        ):
            kpi.addWidget(label)
        kpi.addStretch()
        lay.addLayout(kpi)

        self._product_table = DataTable(_PRODUCT_HEADERS)
        self._product_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._product_table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeMode.Stretch)
        self._product_table.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeMode.Stretch)
        self._product_table.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        lay.addWidget(self._product_table, stretch=3)

        self._category_table = DataTable(_CATEGORY_HEADERS)
        self._category_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._category_table.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        lay.addWidget(self._category_table, stretch=1)

        self._doc_table = DataTable(_DOC_HEADERS)
        self._doc_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self._doc_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self._doc_table.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        lay.addWidget(self._doc_table, stretch=2)

        self._anomaly_label = QLabel("Problemfaelle:")
        lay.addWidget(self._anomaly_label)
        self._anomaly_table = DataTable(["Hinweis"])
        self._anomaly_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._anomaly_table.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        lay.addWidget(self._anomaly_table, stretch=1)

        return page

    def _on_period_changed(self) -> None:
        is_custom = self._period_combo.currentData() == "custom"
        self._custom_start.setEnabled(bool(is_custom))
        self._custom_end.setEnabled(bool(is_custom))

    def _run_musikheroes(self, *, use_cache: bool) -> None:
        if self._commission_worker is not None and self._commission_worker.isRunning():
            return
        commission = self._commission_service
        if not self._active_profile_key:
            QMessageBox.warning(self, "Provisionen", "Kein Abrechnungsprofil ausgewaehlt.")
            return

        period_key = str(self._period_combo.currentData() or "last_month")
        reference = self._reference_date.date().toPython()
        custom_start = self._custom_start.date().toPython() if period_key == "custom" else None
        custom_end = self._custom_end.date().toPython() if period_key == "custom" else None

        try:
            period = commission.resolve_period(
                period_key,
                reference_date=reference,
                custom_start=custom_start,
                custom_end=custom_end,
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Provisionen", str(exc))
            return

        profile_label = commission.get_profile(self._active_profile_key).label
        self._commission_status.setText(f"{profile_label}-Abrechnung wird geladen...")
        profile_key = self._active_profile_key
        include_cancellations = self._include_cancellations.isChecked()
        include_credit_notes = self._include_credit_notes.isChecked()

        def job() -> CommissionRunResult:
            return commission.run_profile(
                profile_key,
                period,
                include_cancellation_invoices=include_cancellations,
                include_credit_notes=include_credit_notes,
                refresh_data=not use_cache,
            )

        self._commission_worker = BackgroundWorker(job)
        self._commission_worker.signals.result.connect(self._on_musikheroes_loaded)
        self._commission_worker.signals.error.connect(self._on_error)
        self._commission_worker.start()

    def _on_musikheroes_loaded(self, payload: object) -> None:
        if not isinstance(payload, CommissionRunResult):
            return
        self._last_commission_result = payload

        period = payload.period
        self._commission_status.setText(
            f"{payload.profile.label}: {period.start.isoformat()} bis {period.end.isoformat()}"
        )

        self._kpi_qty.setText(f"Menge: {payload.summary.total_net_quantity:.2f}")
        self._kpi_net.setText(f"Netto: {payload.summary.total_net_amount:.2f} EUR")
        self._kpi_gross.setText(f"Brutto: {payload.summary.total_gross_amount:.2f} EUR")
        self._kpi_corr.setText(f"Korrekturen: {payload.summary.total_correction_quantity:.2f}")
        self._kpi_docs.setText(f"Belege: {payload.summary.document_count}")
        self._kpi_anomalies.setText(f"Problemfaelle: {payload.summary.anomaly_count}")

        self._populate_product_table(payload)
        self._populate_category_table(payload)
        self._populate_doc_table(payload)
        self._populate_anomaly_table(payload)

    def _export_commission_csv(self) -> None:
        result = self._last_commission_result
        if result is None:
            QMessageBox.information(self, "Provisionen", "Bitte zuerst eine Abrechnung laden.")
            return
        if self._export_worker is not None and self._export_worker.isRunning():
            return

        default_name = f"provision_{result.profile.key}_{result.period.start.isoformat()}_{result.period.end.isoformat()}.csv"
        path, _ = QFileDialog.getSaveFileName(self, "CSV exportieren", default_name, "CSV (*.csv)")
        if not path:
            return

        def job() -> str:
            rows = [
                [
                    "sku",
                    "name",
                    "sold_quantity",
                    "canceled_quantity",
                    "credited_quantity",
                    "net_quantity",
                    "net_amount",
                    "gross_amount",
                    "categories",
                    "warning",
                ]
            ]
            for row in result.product_rows:
                rows.append(
                    [
                        row.sku,
                        row.name,
                        f"{row.sold_quantity:.2f}",
                        f"{row.canceled_quantity:.2f}",
                        f"{row.credited_quantity:.2f}",
                        f"{row.net_quantity:.2f}",
                        f"{row.net_amount:.2f}",
                        f"{row.gross_amount:.2f}",
                        ", ".join(row.category_names),
                        row.warning,
                    ]
                )
            with Path(path).open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle, delimiter=";")
                writer.writerows(rows)
            return path

        self._export_worker = BackgroundWorker(job)
        self._export_worker.signals.result.connect(lambda payload: QMessageBox.information(self, "Provisionen", f"CSV exportiert: {payload}"))
        self._export_worker.signals.error.connect(lambda exc: QMessageBox.critical(self, "Provisionen", f"CSV-Export fehlgeschlagen: {exc}"))
        self._export_worker.signals.finished.connect(lambda: setattr(self, "_export_worker", None))
        self._export_worker.start()

    def _export_commission_xlsx(self) -> None:
        result = self._last_commission_result
        if result is None:
            QMessageBox.information(self, "Provisionen", "Bitte zuerst eine Abrechnung laden.")
            return
        if self._export_worker is not None and self._export_worker.isRunning():
            return

        default_name = f"provision_{result.profile.key}_{result.period.start.isoformat()}_{result.period.end.isoformat()}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "XLSX exportieren", default_name, "Excel (*.xlsx)")
        if not path:
            return

        def job() -> str:
            wb = Workbook()
            ws_products = wb.active
            ws_products.title = "Produkte"
            ws_products.append(
                [
                    "SKU",
                    "Name",
                    "Verkauft",
                    "Storno",
                    "Gutschrift",
                    "Netto-Menge",
                    "Netto EUR",
                    "Brutto EUR",
                    "Kategorien",
                    "Warnung",
                ]
            )
            for row in result.product_rows:
                ws_products.append(
                    [
                        row.sku,
                        row.name,
                        row.sold_quantity,
                        row.canceled_quantity,
                        row.credited_quantity,
                        row.net_quantity,
                        row.net_amount,
                        row.gross_amount,
                        ", ".join(row.category_names),
                        row.warning,
                    ]
                )

            ws_categories = wb.create_sheet("Kategorien")
            ws_categories.append(["Kategorie", "Menge", "Netto EUR", "Brutto EUR", "Anteil Netto %"])
            for row in result.category_rows:
                ws_categories.append(
                    [
                        row.category_name,
                        row.quantity,
                        row.net_amount,
                        row.gross_amount,
                        row.share_of_net_amount * 100.0,
                    ]
                )

            ws_docs = wb.create_sheet("Belege")
            ws_docs.append(["Beleg", "Datum", "Typ", "SKU", "Menge", "Netto", "Regel", "Warnung"])
            for row in result.document_rows:
                ws_docs.append(
                    [
                        row.document_number,
                        row.document_date,
                        row.document_type,
                        row.sku,
                        row.signed_quantity,
                        row.signed_net,
                        row.rule,
                        row.warning,
                    ]
                )

            ws_anomalies = wb.create_sheet("Problemfaelle")
            ws_anomalies.append(["Hinweis"])
            for warning in result.anomalies:
                ws_anomalies.append([warning])
            wb.save(path)
            return path

        self._export_worker = BackgroundWorker(job)
        self._export_worker.signals.result.connect(lambda payload: QMessageBox.information(self, "Provisionen", f"XLSX exportiert: {payload}"))
        self._export_worker.signals.error.connect(lambda exc: QMessageBox.critical(self, "Provisionen", f"XLSX-Export fehlgeschlagen: {exc}"))
        self._export_worker.signals.finished.connect(lambda: setattr(self, "_export_worker", None))
        self._export_worker.start()

    def _copy_commission_summary(self) -> None:
        result = self._last_commission_result
        if result is None:
            QMessageBox.information(self, "Provisionen", "Bitte zuerst eine Abrechnung laden.")
            return

        buffer = StringIO()
        writer = csv.writer(buffer, delimiter=";", lineterminator="\n")
        writer.writerow([f"Profil: {result.profile.label}"])
        writer.writerow([f"Zeitraum: {result.period.start.isoformat()} bis {result.period.end.isoformat()}"])
        writer.writerow([])
        writer.writerow(["SKU", "Name", "Netto-Menge", "Netto EUR", "Brutto EUR"])
        for row in result.product_rows:
            writer.writerow(
                [
                    row.sku,
                    row.name,
                    f"{row.net_quantity:.2f}",
                    f"{row.net_amount:.2f}",
                    f"{row.gross_amount:.2f}",
                ]
            )

        clipboard = QApplication.clipboard()
        clipboard.setText(buffer.getvalue())
        QMessageBox.information(self, "Provisionen", "Abrechnung in die Zwischenablage kopiert.")

    def _populate_product_table(self, result: CommissionRunResult) -> None:
        self._product_table.set_data(
            [
                {
                    "SKU": row.sku,
                    "Name": row.name,
                    "Verkauft": f"{row.sold_quantity:.2f}",
                    "Storno": f"{row.canceled_quantity:.2f}",
                    "Gutschrift": f"{row.credited_quantity:.2f}",
                    "Netto-Menge": f"{row.net_quantity:.2f}",
                    "Netto EUR": f"{row.net_amount:.2f}",
                    "Brutto EUR": f"{row.gross_amount:.2f}",
                    "Kategorien": ", ".join(row.category_names),
                    "Warnung": row.warning,
                    "__align__Verkauft": "right",
                    "__align__Storno": "right",
                    "__align__Gutschrift": "right",
                    "__align__Netto-Menge": "right",
                    "__align__Netto EUR": "right",
                    "__align__Brutto EUR": "right",
                }
                for row in result.product_rows
            ]
        )

    def _populate_category_table(self, result: CommissionRunResult) -> None:
        self._category_table.set_data(
            [
                {
                    "Kategorie": row.category_name,
                    "Menge": f"{row.quantity:.2f}",
                    "Netto EUR": f"{row.net_amount:.2f}",
                    "Brutto EUR": f"{row.gross_amount:.2f}",
                    "Anteil Netto": f"{row.share_of_net_amount * 100.0:.2f} %",
                    "__align__Menge": "right",
                    "__align__Netto EUR": "right",
                    "__align__Brutto EUR": "right",
                    "__align__Anteil Netto": "right",
                }
                for row in result.category_rows
            ]
        )

    def _populate_doc_table(self, result: CommissionRunResult) -> None:
        self._doc_table.set_data(
            [
                {
                    "Beleg": item.document_number,
                    "Datum": item.document_date,
                    "Typ": item.document_type,
                    "SKU": item.sku,
                    "Menge": f"{item.signed_quantity:.2f}",
                    "Netto": f"{item.signed_net:.2f}",
                    "Regel": item.rule,
                    "__align__Menge": "right",
                    "__align__Netto": "right",
                }
                for item in result.document_rows
            ]
        )

    def _populate_anomaly_table(self, result: CommissionRunResult) -> None:
        show = self._show_anomalies.isChecked()
        self._anomaly_label.setVisible(show)
        self._anomaly_table.setVisible(show)
        if not show:
            return

        self._anomaly_table.set_data([{"Hinweis": warning} for warning in result.anomalies])

    # ------------------------------------------------------------------
    # Legacy page: article list with computed royalties
    # ------------------------------------------------------------------

    def _build_articles_tab(self) -> QWidget:
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(8)

        bar = QHBoxLayout()
        self._art_status = QLabel("Artikelliste laden...")
        bar.addWidget(self._art_status)
        bar.addStretch()
        refresh_btn = QPushButton("Aktualisieren")
        refresh_btn.clicked.connect(self._load_articles)
        bar.addWidget(refresh_btn)
        lay.addLayout(bar)

        self._art_table = DataTable(_ARTICLE_HEADERS)
        self._art_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._art_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)
        self._art_table.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        self._art_table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        lay.addWidget(self._art_table)

        info = QLabel("Artikelliste in DB: Einstellungen > Schluessel-Verwaltung > calculation.articles (JSON-Array).")
        info.setObjectName("infoLabel")
        info.setWordWrap(True)
        lay.addWidget(info)
        return page

    def _load_articles(self) -> None:
        svc: CalculationService = self._container.resolve(CalculationService)

        def job() -> list[ArticleEntry]:
            return svc.load_articles()

        self._worker = BackgroundWorker(job)
        self._worker.signals.result.connect(self._on_articles_loaded)
        self._worker.signals.error.connect(self._on_error)
        self._worker.start()

    def _on_articles_loaded(self, rows: object) -> None:
        if not isinstance(rows, list):
            return
        self._articles = rows  # type: ignore[assignment]
        if not self._articles:
            self._art_status.setText("Keine Artikel — bitte calculation.articles in Einstellungen befuellen.")
        else:
            self._art_status.setText(f"{len(self._articles)} Artikel geladen")
        self._populate_articles(self._articles)

    def _populate_articles(self, items: list[ArticleEntry]) -> None:
        payload: list[dict[str, object]] = []
        for art in items:
            res = calculate_royalty(art.gross_price, vat_pct=art.vat_pct, royalty_pct=art.royalty_pct)
            payload.append(
                {
                    "Titel": art.title,
                    "Brutto EUR": f"{art.gross_price:.2f}",
                    "MwSt %": f"{art.vat_pct:.1f}",
                    "Provision %": f"{art.royalty_pct:.1f}",
                    "Netto EUR": f"{res.net:.2f}",
                    "MwSt EUR": f"{res.vat_amount:.2f}",
                    "Provision EUR": f"{res.royalty_amount:.2f}",
                    "Notiz": art.note,
                    "__align__Brutto EUR": "right",
                    "__align__MwSt %": "center",
                    "__align__Provision %": "center",
                    "__align__Netto EUR": "right",
                    "__align__MwSt EUR": "right",
                    "__align__Provision EUR": "right",
                }
            )
        self._art_table.set_data(payload)

    # ------------------------------------------------------------------
    # Legacy page: quick calculator
    # ------------------------------------------------------------------

    def _build_calc_tab(self) -> QWidget:
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(12)

        grp = QGroupBox("Eingabe")
        form = QFormLayout(grp)

        self._calc_gross = QDoubleSpinBox()
        self._calc_gross.setRange(0.0, 99999.99)
        self._calc_gross.setDecimals(2)
        self._calc_gross.setSuffix(" EUR")
        self._calc_gross.setValue(10.00)
        form.addRow("Bruttopreis:", self._calc_gross)

        self._calc_vat = QDoubleSpinBox()
        self._calc_vat.setRange(0.0, 100.0)
        self._calc_vat.setDecimals(1)
        self._calc_vat.setSuffix(" %")
        self._calc_vat.setValue(10.0)
        form.addRow("MwSt-Satz:", self._calc_vat)

        self._calc_royalty = QDoubleSpinBox()
        self._calc_royalty.setRange(0.0, 100.0)
        self._calc_royalty.setDecimals(2)
        self._calc_royalty.setSuffix(" %")
        self._calc_royalty.setValue(0.0)
        form.addRow("Provisionssatz (auf Netto):", self._calc_royalty)

        lay.addWidget(grp)

        calc_btn = QPushButton("Berechnen")
        calc_btn.clicked.connect(self._run_calc)
        lay.addWidget(calc_btn)

        res_grp = QGroupBox("Ergebnis")
        res_lay = QFormLayout(res_grp)

        self._res_net = QLabel("—")
        res_lay.addRow("Nettobetrag:", self._res_net)
        self._res_vat = QLabel("—")
        res_lay.addRow("MwSt-Betrag:", self._res_vat)
        self._res_provision = QLabel("—")
        res_lay.addRow("Provision:", self._res_provision)
        self._res_net_after = QLabel("—")
        res_lay.addRow("Netto nach Provision:", self._res_net_after)
        lay.addWidget(res_grp)

        lay.addStretch()
        return page

    def _run_calc(self) -> None:
        gross = self._calc_gross.value()
        vat = self._calc_vat.value()
        prov = self._calc_royalty.value()
        res = calculate_royalty(gross, vat_pct=vat, royalty_pct=prov)
        self._res_net.setText(f"{res.net:.4f} EUR")
        self._res_vat.setText(f"{res.vat_amount:.4f} EUR")
        self._res_provision.setText(f"{res.royalty_amount:.4f} EUR")
        self._res_net_after.setText(f"{res.net_after_royalty:.4f} EUR")

    def _on_error(self, exc: BaseException) -> None:
        logger.exception("CalculationView error: %s", exc)
        QMessageBox.critical(self, "Fehler", str(exc))
